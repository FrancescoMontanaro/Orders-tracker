import csv
import asyncio
import zipfile
import openpyxl
from pathlib import Path
from typing import Optional, AsyncGenerator
from datetime import date, datetime, timezone, timedelta

from sqlalchemy import select, func, asc, desc
from sqlalchemy.ext.asyncio import AsyncSession

from ....core.config import settings
from ....db.session import db_session
from .exceptions import JobAlreadyExistsException
from ....core.ws_manager import export_ws_manager
from ....models import Pagination, ListingQueryParams
from ....db.orm.notification import NotificationTypeEnum
from .models import ExportJobStart, ExportReportJobStart, ExportJob
from ..notifications.service import create_notification as create_notification_service
from ....db.orm.export_job import ExportJobORM, ExportStatusEnum, ExportFormatEnum, ExportEntityEnum
from .constants import ALLOWED_EXPORT_JOBS_SORTING_FIELDS, ENTITY_HEADERS, ENTITY_LABELS, ENTITY_SHEET_NAMES
from .utils import (
    iter_customers,
    iter_products,
    iter_orders,
    iter_order_items,
    iter_expenses,
    iter_incomes,
    iter_lots,
    iter_notes
)
from ..reports.models import (
    ProductSalesRequest,
    ExpensesCategoriesRequest,
    IncomeCategoriesRequest,
    CustomerSalesRequest,
    CashflowRequest
)
from ..reports.service import (
    report_product_sales as report_product_sales_service,
    report_expenses_categories as report_expenses_categories_service,
    report_income_categories as report_income_categories_service,
    report_customer_sales as report_customer_sales_service,
    report_cashflow as report_cashflow_service
)

# ====================== #
# ===== Public API ===== #
# ====================== #

async def create_export_job(payload: ExportJobStart, user_id: int) -> ExportJobORM:
    """
    Persist a new export job (status = PENDING) and return it.

    Raises HTTP 409 if any of the requested entity_types already has a pending or
    running job for this user, to avoid duplicate concurrent exports.

    Parameters:
    - payload (ExportJobStart): The export job parameters.
    - user_id (int): The ID of the user requesting the export.

    Returns:
    - ExportJobORM: The newly created export job.
    """

    # Serialise entity_types to a plain list of strings for JSON storage
    entity_types_values = [e.value for e in payload.entity_types]

    # Create a new database session for this operation
    async with db_session() as session:
        # Reject if any active job belonging to this user already covers one or
        # more of the requested entities, to prevent duplicate concurrent exports
        active_jobs = (
            await session.execute(
                select(ExportJobORM).where(
                    ExportJobORM.user_id == user_id,
                    ExportJobORM.status.in_([ExportStatusEnum.PENDING, ExportStatusEnum.RUNNING])
                )
            )
        ).scalars().all()

        # Check for overlap between requested entities and active jobs' entities.
        for active in active_jobs:
            overlap = set(active.entity_types) & set(entity_types_values)
            if overlap:
                raise JobAlreadyExistsException()

        # Create and persist the new job
        job = ExportJobORM(
            user_id = user_id,
            entity_types = entity_types_values,
            format = payload.format,
            start_date = payload.start_date,
            end_date = payload.end_date
        )

        # Persist the job to the database
        session.add(job)

        # Commit the transaction to generate the job ID and make it available for the background task
        await session.commit()
        await session.refresh(job)

        # Notify connected tabs in real time that a new job is pending
        job_data = ExportJob.model_validate(job).model_dump(mode='json')
        await export_ws_manager.broadcast(user_id, job_data)

        # Return the newly created job
        return job


async def create_report_export_job(payload: ExportReportJobStart, user_id: int) -> ExportJobORM:
    """
    Persist a new report export job (status = PENDING) and return it.

    Raises JobAlreadyExistsException if an active job for the same report type
    and user already exists, to avoid duplicate concurrent exports.

    Parameters:
    - payload (ExportReportJobStart): The report export job parameters.
    - user_id (int): The ID of the user requesting the export.

    Returns:
    - ExportJobORM: The newly created export job.
    """

    # The entity type for this job is always the single report type
    entity_types_values = [payload.report_type.value]

    # Build the report-specific parameters dict to persist alongside the job
    report_params: dict = {"include_incomes": payload.include_incomes}
    if payload.product_ids:
        report_params["product_ids"] = payload.product_ids
    if payload.expense_category_ids:
        report_params["expense_category_ids"] = payload.expense_category_ids
    if payload.income_category_ids:
        report_params["income_category_ids"] = payload.income_category_ids
    if payload.customer_id:
        report_params["customer_id"] = payload.customer_id

    # Create a new database session for this operation
    async with db_session() as session:
        # Reject if an active job for the same report type already exists for this user
        active_jobs = (
            await session.execute(
                select(ExportJobORM).where(
                    ExportJobORM.user_id == user_id,
                    ExportJobORM.status.in_([ExportStatusEnum.PENDING, ExportStatusEnum.RUNNING])
                )
            )
        ).scalars().all()

        # Check for overlap between requested entity and active jobs' entities
        for active in active_jobs:
            overlap = set(active.entity_types) & set(entity_types_values)
            if overlap:
                raise JobAlreadyExistsException()

        # Create and persist the new job
        job = ExportJobORM(
            user_id = user_id,
            entity_types = entity_types_values,
            format = payload.format,
            start_date = payload.start_date,
            end_date = payload.end_date,
            report_params = report_params
        )

        # Persist the job to the database
        session.add(job)

        # Commit the transaction to generate the job ID and make it available for the background task
        await session.commit()
        await session.refresh(job)

        # Notify connected tabs in real time that a new job is pending
        job_data = ExportJob.model_validate(job).model_dump(mode='json')
        await export_ws_manager.broadcast(user_id, job_data)

        # Return the newly created job
        return job


async def list_export_jobs(params: ListingQueryParams, user_id: int) -> Pagination[ExportJob]:
    """
    List export jobs for the given user with pagination, filtering and sorting.

    Parameters:
    - params (ListingQueryParams): The query parameters for listing jobs.
    - user_id (int): The owner's user id (always applied as a security filter).

    Returns:
    - Pagination[ExportJob]: Paginated list of export jobs.
    """

    # Compute pagination params
    page = max(1, params.page)
    size = params.size
    offset = (page - 1) * size

    # Create a new database session for this operation
    async with db_session() as session:
        # Build filter conditions
        filters = params.filters or {}
        conditions = [ExportJobORM.user_id == user_id]

        # Only apply filters for allowed fields and non-null values
        for field, value in filters.items():
            # Ignore invalid fields or null values
            if value is None or field not in ALLOWED_EXPORT_JOBS_SORTING_FIELDS:
                continue

            # Map the field name to the actual ORM column
            col = ALLOWED_EXPORT_JOBS_SORTING_FIELDS[field]
            conditions.append(col == value)

        # Base statement
        stmt = select(ExportJobORM).where(*conditions)

        # Count total
        count_stmt = select(func.count()).select_from(stmt.subquery())
        total = (await session.execute(count_stmt)).scalar_one()

        # Apply sorting
        if params.sort:
            # Only apply sorting for allowed fields, ignore invalid ones
            for sort_param in params.sort:
                if sort_param.field in ALLOWED_EXPORT_JOBS_SORTING_FIELDS:
                    col = ALLOWED_EXPORT_JOBS_SORTING_FIELDS[sort_param.field]
                    stmt = stmt.order_by(asc(col) if sort_param.order == "asc" else desc(col))
        else:
            # Default sorting by creation date desc
            stmt = stmt.order_by(desc(ExportJobORM.created_at))

        # Apply pagination
        stmt = stmt.offset(offset).limit(size)

        # Execute the query and map results to Pydantic models
        rows = (await session.execute(stmt)).scalars().all()
        items = [ExportJob.model_validate(row) for row in rows]

        # Return paginated response
        return Pagination(total=total, items=items)


async def get_export_job(job_id: int, user_id: int) -> ExportJobORM | None:
    """
    Retrieve a job only if it belongs to the given user.
    Returns None on not found or ownership mismatch.

    Parameters:
    - job_id (int): The ID of the export job to retrieve.
    - user_id (int): The owner's user id (always applied as a security filter).

    Returns:
    - ExportJobORM | None: The export job if found and owned by the user, else None.
    """

    # Create a new database session for this operation
    async with db_session() as session:
        # Retrieve the job by ID
        job = await session.get(ExportJobORM, job_id)

        # Return the job only if it exists and belongs to the user, else return None
        if not job or job.user_id != user_id:
            return None
        
        # If the job exists and belongs to the user, return it
        return job


async def run_export_job(job_id: int, exports_dir: str) -> None:
    """
    Background task: generate the export file and update job lifecycle fields.

    Flow:
        PENDING → RUNNING → COMPLETED (file written to disk)
                          → FAILED (exception message stored)

    Notes:
        - The list of entities to export is stored as a JSON array of strings in the database.
        - When ORDERS is selected, ORDER_ITEMS is automatically appended (the two tables are always exported together).
        - Multiple entities always produce either a ZIP (CSV) or a multi-sheet XLSX, never a bare CSV.
        - Synchronous file I/O (csv / openpyxl) is offloaded to a thread via asyncio.to_thread so the event loop is never blocked.

    Parameters:
    - job_id (int): The ID of the export job to run.
    - exports_dir (str): The directory where export files should be saved.
    """

    # Create a new database session for this operation
    async with db_session() as session:
        # Retrieve the job by ID
        job = await session.get(ExportJobORM, job_id)

        # If the job doesn't exist (should not happen), just return and do nothing
        if not job:
            return

        # Mark the job as running
        job.status = ExportStatusEnum.RUNNING
        job.started_at = datetime.now(timezone.utc)
        await session.commit()

        # Refresh to reload attributes expired by the commit
        await session.refresh(job)

        # Capture scalar values before any further commit expires the object again
        job_user_id = job.user_id
        job_id_val = job.id
        # Resolve entity enums from the stored JSON strings
        job_entities = [ExportEntityEnum(v) for v in job.entity_types]
        # Build a human-readable label summarising all selected entities
        job_entity_label = ", ".join(ENTITY_LABELS.get(e, e.value) for e in job_entities)

        # Broadcast RUNNING status to connected tabs
        job_data = ExportJob.model_validate(job).model_dump(mode='json')
        await export_ws_manager.broadcast(job_user_id, job_data)

        try:
            # Run the export with a timeout to prevent runaway jobs
            async with asyncio.timeout(settings.export_job_timeout_seconds):
                # Determine the entities to export directly from the stored list.
                # ORDER_ITEMS is automatically appended when ORDERS is selected,
                # since the two tables are tightly coupled and always exported together.
                entities: list[ExportEntityEnum] = []
                for e in job_entities:
                    entities.append(e)
                    if e == ExportEntityEnum.ORDERS and ExportEntityEnum.ORDER_ITEMS not in job_entities:
                        entities.append(ExportEntityEnum.ORDER_ITEMS)

                # Build the output file path and ensure the directory exists
                out_dir = Path(exports_dir)
                out_dir.mkdir(parents=True, exist_ok=True)

                # Detect whether this is a report-based job (all entities start with "report_")
                is_report_job = all(e.value.startswith("report_") for e in entities)

                if is_report_job:
                    # Report export: each entity maps to a report function; always XLSX
                    # (CSV is also supported for single-report jobs via a ZIP for multi-report)
                    report_params = job.report_params or {}
                    is_multi_report = len(entities) > 1
                    if is_multi_report and job.format == ExportFormatEnum.CSV:
                        file_path = str(out_dir / f"{job_id}.zip")
                        await _build_report_zip_csv(file_path, entities, job.start_date, job.end_date, report_params)
                    elif is_multi_report or job.format == ExportFormatEnum.XLSX:
                        file_path = str(out_dir / f"{job_id}.xlsx")
                        await _build_report_xlsx(file_path, entities, job.start_date, job.end_date, report_params)
                    else:
                        file_path = str(out_dir / f"{job_id}.csv")
                        await _build_report_csv(file_path, entities[0], job.start_date, job.end_date, report_params)

                else:
                    # Standard table export
                    is_multi = len(entities) > 1
                    if is_multi and job.format == ExportFormatEnum.CSV:
                        file_path = str(out_dir / f"{job_id}.zip")
                        await _build_zip_csv(file_path, entities, job.start_date, job.end_date, session)
                    elif is_multi or job.format == ExportFormatEnum.XLSX:
                        file_path = str(out_dir / f"{job_id}.xlsx")
                        await _build_xlsx(file_path, entities, job.start_date, job.end_date, session)
                    else:
                        file_path = str(out_dir / f"{job_id}.csv")
                        await _build_csv(file_path, entities[0], job.start_date, job.end_date, session)

            # Mark the job as completed
            now = datetime.now(timezone.utc)
            job.status = ExportStatusEnum.COMPLETED
            job.completed_at = now
            job.file_path = file_path
            await session.commit()
            await session.refresh(job)

            # Broadcast COMPLETED status to connected tabs
            job_data = ExportJob.model_validate(job).model_dump(mode='json')
            await export_ws_manager.broadcast(job_user_id, job_data)

            # Notify the user that the export is ready
            await create_notification_service(
                user_id   = job_user_id,
                type      = NotificationTypeEnum.EXPORT_COMPLETED,
                title     = "Export completato",
                message   = f"{job_entity_label} - il file è pronto, puoi scaricarlo dalla pagina Export.",
                entity_id = job_id_val,
            )

        except TimeoutError:
            # Job exceeded the configured timeout
            job.status = ExportStatusEnum.FAILED
            job.error_message = f"Job exceeded the maximum allowed duration of {settings.export_job_timeout_seconds}s"
            await session.commit()
            await session.refresh(job)

            # Broadcast FAILED status to connected tabs
            job_data = ExportJob.model_validate(job).model_dump(mode='json')
            await export_ws_manager.broadcast(job_user_id, job_data)

            # Notify the user that the export failed
            await create_notification_service(
                user_id   = job_user_id,
                type      = NotificationTypeEnum.EXPORT_FAILED,
                title     = "Export fallito",
                message   = f"{job_entity_label} - l'elaborazione ha impiegato troppo tempo ed è stata interrotta automaticamente.",
                entity_id = job_id_val,
            )

        except Exception as exc:
            # Mark the job as failed and store the error message
            job.status = ExportStatusEnum.FAILED
            job.error_message = str(exc)

            # Commit the transaction to persist the failure status and error message before sending the notification
            await session.commit()
            await session.refresh(job)

            # Broadcast FAILED status to connected tabs
            job_data = ExportJob.model_validate(job).model_dump(mode='json')
            await export_ws_manager.broadcast(job_user_id, job_data)

            # Notify the user that the export failed
            await create_notification_service(
                user_id   = job_user_id,
                type      = NotificationTypeEnum.EXPORT_FAILED,
                title     = "Export fallito",
                message   = f"{job_entity_label} - si è verificato un errore durante l'elaborazione. Riprova o contatta l'assistenza.",
                entity_id = job_id_val,
            )

            # Re-raise the exception to allow logging and monitoring systems to capture it, 
            # while the job status and notification have already been handled
            raise


async def expire_stale_export_jobs(timeout_seconds: int) -> None:
    """
    Mark as FAILED any PENDING or RUNNING job that has been started more than
    timeout_seconds ago without reaching a terminal state.

    Called periodically by the application watchdog (see context_manager.py).

    Parameters:
    - timeout_seconds (int): Maximum allowed duration for a single export job.
    """

    # Compute the cutoff time: jobs started before this moment are considered stale
    cutoff = datetime.now(timezone.utc) - timedelta(seconds=timeout_seconds)

    # Create a new database session for this operation
    async with db_session() as session:
        # Find jobs that started before the cutoff and are still in a non-terminal state
        stmt = select(ExportJobORM).where(
            ExportJobORM.status.in_([ExportStatusEnum.PENDING, ExportStatusEnum.RUNNING]),
            ExportJobORM.started_at <= cutoff
        )

        # Execute the query and get the list of stale jobs
        stale_jobs = (await session.execute(stmt)).scalars().all()

        # Mark each stale job as failed
        for job in stale_jobs:
            job.status = ExportStatusEnum.FAILED
            job.error_message = f"Job exceeded the maximum allowed duration of {timeout_seconds}s and was automatically terminated"

        # Persist the changes if there are any stale jobs
        if stale_jobs:
            await session.commit()


# ========================= #
# ===== Data fetching ===== #
# ========================= #


async def _iter_entity(
    entity: ExportEntityEnum,
    start_date: Optional[date],
    end_date: Optional[date],
    session: AsyncSession
) -> AsyncGenerator[list[list], None]:
    """
    Dispatch async generator: yields row batches for the given entity type.

    Parameters:
    - entity (ExportEntityEnum): The entity to stream.
    - start_date (Optional[date]): Optional start date filter.
    - end_date (Optional[date]): Optional end date filter.
    - session (AsyncSession): The database session.
    """

    # Dispatch to the correct entity generator based on the entity type.
    match entity:
        case ExportEntityEnum.CUSTOMERS:
            async for batch in iter_customers(session):
                yield batch
        case ExportEntityEnum.PRODUCTS:
            async for batch in iter_products(session):
                yield batch
        case ExportEntityEnum.ORDERS:
            async for batch in iter_orders(session, start_date, end_date):
                yield batch
        case ExportEntityEnum.ORDER_ITEMS:
            async for batch in iter_order_items(session, start_date, end_date):
                yield batch
        case ExportEntityEnum.EXPENSES:
            async for batch in iter_expenses(session, start_date, end_date):
                yield batch
        case ExportEntityEnum.INCOMES:
            async for batch in iter_incomes(session, start_date, end_date):
                yield batch
        case ExportEntityEnum.LOTS:
            async for batch in iter_lots(session, start_date, end_date):
                yield batch
        case ExportEntityEnum.NOTES:
            async for batch in iter_notes(session, start_date, end_date):
                yield batch


# ======================== #
# ===== File writers ===== #
# ======================== #

async def _build_csv(
    file_path: str,
    entity: ExportEntityEnum,
    start_date: Optional[date],
    end_date: Optional[date],
    session: AsyncSession,
) -> None:
    """
    Build a CSV file for a single entity, fetching rows from the database in
    batches to avoid loading the entire dataset into memory at once.

    Parameters:
    - file_path (str): The output file path.
    - entity (ExportEntityEnum): The entity to export.
    - start_date (Optional[date]): Optional start date filter.
    - end_date (Optional[date]): Optional end date filter.
    - session (AsyncSession): The database session.
    """

    # Open the output file in write mode with UTF-8 encoding and create a CSV writer
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        # Create a CSV writer that will write rows to the file
        writer = csv.writer(f)

        # Write the header row
        writer.writerow(ENTITY_HEADERS[entity])

        # Fetch and write rows in batches to avoid loading all data into memory at once
        async for batch in _iter_entity(entity, start_date, end_date, session):
            writer.writerows(batch)


async def _build_zip_csv(
    file_path: str,
    entities: list[ExportEntityEnum],
    start_date: Optional[date],
    end_date: Optional[date],
    session: AsyncSession,
) -> None:
    """
    Build a ZIP archive containing one CSV file per entity, fetching rows from
    the database in batches to avoid loading the entire dataset into memory at once.

    Parameters:
    - file_path (str): The output ZIP file path.
    - entities (list[ExportEntityEnum]): The entities to export, one CSV per entity.
    - start_date (Optional[date]): Optional start date filter.
    - end_date (Optional[date]): Optional end date filter.
    - session (AsyncSession): The database session.
    """

    out_dir = Path(file_path).parent
    job_stem = Path(file_path).stem
    tmp_files: list[tuple[ExportEntityEnum, str]] = []

    try:
        # Write a temporary CSV file for each entity
        for entity in entities:
            tmp_path = str(out_dir / f"_tmp_{job_stem}_{entity.value}.csv")
            await _build_csv(tmp_path, entity, start_date, end_date, session)
            tmp_files.append((entity, tmp_path))

        # Pack all CSV files into a single ZIP archive
        def _write_zip() -> None:
            with zipfile.ZipFile(file_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for entity, tmp_path in tmp_files:
                    zf.write(tmp_path, arcname=f"{ENTITY_SHEET_NAMES[entity]}.csv")

        await asyncio.to_thread(_write_zip)

    finally:
        # Remove temporary CSV files regardless of outcome
        for _, tmp_path in tmp_files:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except Exception:
                pass


async def _build_xlsx(
    file_path: str,
    entities: list[ExportEntityEnum],
    start_date: Optional[date],
    end_date: Optional[date],
    session: AsyncSession,
) -> None:
    """
    Build a multi-sheet XLSX file, fetching rows from the database in batches.
    Uses openpyxl write-only mode to reduce memory footprint.

    Parameters:
    - file_path (str): The output file path.
    - entities (list[ExportEntityEnum]): The entities to export, one sheet per entity.
    - start_date (Optional[date]): Optional start date filter.
    - end_date (Optional[date]): Optional end date filter.
    - session (AsyncSession): The database session.
    """

    # Write-only mode streams rows directly through an internal buffer, avoiding
    # the need to hold the full workbook in memory
    wb = openpyxl.Workbook(write_only=True)

    # For each entity, create a new sheet and write rows in batches to avoid loading all data into memory at once
    for entity in entities:
        # Create a new sheet for this entity (openpyxl automatically handles sheet naming conflicts by appending a number)
        ws = wb.create_sheet(title=ENTITY_SHEET_NAMES[entity])

        # Write the header row for this sheet
        ws.append(ENTITY_HEADERS[entity])

        # Fetch and append rows in batches to avoid loading all data into memory at once
        async for batch in _iter_entity(entity, start_date, end_date, session):
            for row in batch:
                ws.append([v if v is not None else "" for v in row])

    # Flush the workbook to disk in a thread to avoid blocking the event loop
    await asyncio.to_thread(wb.save, file_path)


# ================================ #
# ===== Report data helpers  ===== #
# ================================ #

async def _generate_report_rows(
    entity: ExportEntityEnum,
    start_date: Optional[date],
    end_date: Optional[date],
    params: dict,
) -> tuple[list[str], list[list]]:
    """
    Call the appropriate report service function and return (headers, rows).

    The returned rows are flat lists of scalar values ready to be written to
    CSV or XLSX without further transformation.

    Parameters:
    - entity (ExportEntityEnum): One of the REPORT_* entity types.
    - start_date (Optional[date]): Start of the date range (required for all reports).
    - end_date (Optional[date]): End of the date range (required for all reports).
    - params (dict): Report-specific parameters (product_ids, customer_id, etc.).

    Returns:
    - (headers, rows): Column headers and data rows.
    """

    # Get the column headers for this report entity type
    headers = ENTITY_HEADERS[entity]

    # Report exports always require a date range — this is enforced by ExportReportJobStart
    assert start_date is not None and end_date is not None, (
        "start_date and end_date are required for report exports"
    )

    # Call the appropriate report service function based on the entity type and extract rows.
    match entity:
        # For each report type, we call the corresponding service function with the appropriate parameters,
        # then transform the returned data into a list of rows, where each row is a list
        case ExportEntityEnum.REPORT_PRODUCT_SALES:
            data = await report_product_sales_service(
                ProductSalesRequest(
                    start_date = start_date,
                    end_date = end_date,
                    product_ids = params.get("product_ids"),
                )
            )
            rows = [
                [r.product_id, r.product_name, r.total_qty, r.unit, r.revenue]
                for r in data
            ]

        case ExportEntityEnum.REPORT_EXPENSES:
            data = await report_expenses_categories_service(
                ExpensesCategoriesRequest(
                    start_date = start_date,
                    end_date = end_date,
                    category_ids = params.get("expense_category_ids"),
                )
            )
            rows = [
                [r.category_id, r.category_descr, r.amount, r.count]
                for r in data
            ]

        case ExportEntityEnum.REPORT_INCOMES:
            data = await report_income_categories_service(
                IncomeCategoriesRequest(
                    start_date = start_date,
                    end_date = end_date,
                    category_ids = params.get("income_category_ids"),
                )
            )
            rows = [
                [r.category_id, r.category_descr, r.amount, r.count]
                for r in data
            ]

        case ExportEntityEnum.REPORT_CUSTOMER_SALES:
            data = await report_customer_sales_service(
                CustomerSalesRequest(
                    start_date = start_date,
                    end_date = end_date,
                    customer_id = params.get("customer_id", 0),
                )
            )
            # Flatten: repeat customer info on every product row for CSV/XLSX readability
            rows = [
                [
                    data.customer_name,
                    data.total_revenue,
                    r.product_id,
                    r.product_name,
                    r.total_qty,
                    r.unit,
                    r.avg_discount,
                    r.revenue,
                ]
                for r in data.per_product
            ]

        case ExportEntityEnum.REPORT_CASHFLOW:
            data = await report_cashflow_service(
                CashflowRequest(
                    start_date = start_date,
                    end_date = end_date,
                    include_incomes = params.get("include_incomes", True),
                )
            )
            rows = []
            for e in data.entries:
                rows.append(["Ordine", e.order_id, str(e.date), e.amount, ""])
            for inc in data.incomes:
                rows.append(["Entrata extra", inc.id, str(inc.date), inc.amount, inc.note or ""])
            for exp in data.expenses:
                rows.append(["Uscita", exp.id, str(exp.date), exp.amount, exp.note or ""])
            # Summary rows
            rows.append(["", "", "", "", ""])
            rows.append(["TOTALE ENTRATE", "", "", data.entries_total, ""])
            rows.append(["TOTALE USCITE", "", "", data.expenses_total, ""])
            rows.append(["SALDO NETTO", "", "", data.net, ""])

        case _:
            # This should never happen because the API only allows REPORT_* entity types for report export jobs,
            raise ValueError(f"Unknown report entity type: {entity}")

    # Return the column headers and the list of data rows for this report entity type
    return headers, rows


async def _build_report_csv(
    file_path: str,
    entity: ExportEntityEnum,
    start_date: Optional[date],
    end_date: Optional[date],
    params: dict,
) -> None:
    """
    Build a single CSV file for a report entity type.

    Parameters:
    - file_path (str): The output file path.
    - entity (ExportEntityEnum): The REPORT_* entity to export.
    - start_date (Optional[date]): Start of the date range.
    - end_date (Optional[date]): End of the date range.
    - params (dict): Report-specific parameters.
    """

    # Generate the report data rows by calling the appropriate report service function
    headers, rows = await _generate_report_rows(entity, start_date, end_date, params)

    # Write the headers and rows to a CSV file in a thread to avoid blocking the event loop
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


async def _build_report_zip_csv(
    file_path: str,
    entities: list[ExportEntityEnum],
    start_date: Optional[date],
    end_date: Optional[date],
    params: dict,
) -> None:
    """
    Build a ZIP archive containing one CSV file per report entity type.

    Parameters:
    - file_path (str): The output ZIP file path.
    - entities (list[ExportEntityEnum]): The REPORT_* entities to export.
    - start_date (Optional[date]): Start of the date range.
    - end_date (Optional[date]): End of the date range.
    - params (dict): Shared report-specific parameters.
    """

    # Define the output directory and a temporary file path for each entity's CSV, which will be deleted after the ZIP is created
    out_dir = Path(file_path).parent
    job_stem = Path(file_path).stem
    tmp_files: list[tuple[ExportEntityEnum, str]] = []

    try:
        # Write a temporary CSV file for each entity by generating the report data and writing it to disk
        for entity in entities:
            tmp_path = str(out_dir / f"_tmp_{job_stem}_{entity.value}.csv")
            await _build_report_csv(tmp_path, entity, start_date, end_date, params)
            tmp_files.append((entity, tmp_path))

        # Define a synchronous function to create the ZIP archive, since zipfile is not async-aware and we want to offload it to a thread
        def _write_zip() -> None:
            with zipfile.ZipFile(file_path, "w", zipfile.ZIP_DEFLATED) as zf:
                for entity, tmp_path in tmp_files:
                    zf.write(tmp_path, arcname=f"{ENTITY_SHEET_NAMES[entity]}.csv")

        # Pack all CSV files into a single ZIP archive in a thread to avoid blocking the event loop
        await asyncio.to_thread(_write_zip)

    finally:
        # Remove temporary CSV files regardless of outcome to avoid leaving orphaned files on disk
        for _, tmp_path in tmp_files:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except Exception:
                pass


async def _build_report_xlsx(
    file_path: str,
    entities: list[ExportEntityEnum],
    start_date: Optional[date],
    end_date: Optional[date],
    params: dict,
) -> None:
    """
    Build a multi-sheet XLSX file for one or more report entity types.

    Parameters:
    - file_path (str): The output file path.
    - entities (list[ExportEntityEnum]): The REPORT_* entities to export, one sheet each.
    - start_date (Optional[date]): Start of the date range.
    - end_date (Optional[date]): End of the date range.
    - params (dict): Shared report-specific parameters.
    """

    # Write-only mode streams rows directly through an internal buffer, avoiding
    wb = openpyxl.Workbook(write_only=True)

    # For each entity, create a new sheet and write the report data rows, which are generated by calling the appropriate report service function
    for entity in entities:
        # Generate the report data rows for this entity by calling the appropriate report service function
        headers, rows = await _generate_report_rows(entity, start_date, end_date, params)

        # Create a new sheet for this entity (openpyxl automatically handles sheet naming conflicts by appending a number)
        ws = wb.create_sheet(title=ENTITY_SHEET_NAMES[entity])
        ws.append(headers)
        for row in rows:
            ws.append([v if v is not None else "" for v in row])

    # Flush the workbook to disk in a thread to avoid blocking the event loop
    await asyncio.to_thread(wb.save, file_path)